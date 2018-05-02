# -*- coding: UTF-8 -*-
import os
import sys
import pandas as pd
from process import *
from generate import *
from openpyxl import load_workbook

# get file names in current dir, return error message if more than one file
def getFileName(path):
    count = 0

    list = os.listdir(path)
    for i in list:
        if os.path.splitext(i)[1] == '.xlsx':
            count += 1
            if os.path.splitext(i)[0] != 'model':
                infile = i
            else:
                modelFile = i

    if count > 1:
        print u'只能有1个Excel文件，目前目录下有多个！'
        sys.exit(0)
    else:
        return infile


# read all sheets and return a list of df
def readFile(file_name):
    sheets = []
    wb = load_workbook(file_name)
    sheetNames = wb.sheetnames

    for i in range(5):
        try:
            sheetTmp = pd.read_excel(file_name, sheetname=sheetNames[i], header=0, index_col=0)
            sheets.append(sheetTmp)
        except Exception, e:
            print e
            print u'文件读取失败...'

    print u'文件读取成功~~~'
    return [sheets, sheetNames]

# run on command
if __name__ == "__main__":
    path = './'

    inFileName = getFileName(path)
    inSheets, iSheetNames = readFile(inFileName)

    # print inSheets[0]

    objList = [eventProcessor(inSheets[0]), varProcessor(inSheets[1]), pVarProcessor(inSheets[2]),
        eVarProcessor(inSheets[3]), uVarProcessor(inSheets[4])]

    # get output sheets in dataframe
    outSheets = convertToDF(objList)

    # generate excel file
    genFile(outSheets)
